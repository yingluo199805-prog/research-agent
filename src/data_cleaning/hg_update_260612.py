import argparse
import json
import re
import sys
from collections import Counter

import openpyxl


def strip_suffix(name):
    return re.sub(r"\s*\(.*?\)\s*", "", str(name)).strip()


def find_cols(row1):
    cols = {}
    for i, h in enumerate(row1):
        if not h:
            continue
        s = str(h)
        if "品牌-系别" in s:
            cols["brand_series"] = i + 1
        if "品牌-车企划分" in s:
            cols["brand_oem"] = i + 1
        if "一对多" in s:
            cols["multi"] = i + 1
        if "国家" in s:
            cols["country"] = i + 1
    missing = [k for k in ["brand_series", "brand_oem", "multi", "country"] if k not in cols]
    if missing:
        raise RuntimeError(f"missing mapping cols: {missing}")
    return cols


def build_idx(mapping):
    idx = {}
    for k in mapping:
        idx[k] = k
        idx.setdefault(strip_suffix(k), k)
    return idx


def load_maps(main_path):
    wb = openpyxl.load_workbook(main_path, read_only=True, data_only=True)
    ws = wb["海关匹配字段"]
    ws.reset_dimensions()
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    cols = find_cols(row1)

    brand_series = {}
    for row in ws.iter_rows(min_row=3, min_col=cols["brand_series"], max_col=cols["brand_series"] + 1, values_only=True):
        if row[0] and row[1]:
            brand_series[str(row[0]).strip()] = str(row[1]).strip()

    brand_oem = {}
    for row in ws.iter_rows(min_row=3, min_col=cols["brand_oem"], max_col=cols["brand_oem"] + 1, values_only=True):
        if row[0] and row[1]:
            brand_oem[str(row[0]).strip()] = str(row[1]).strip()

    multi_brand_oem = {}
    multi_model_oem = {}
    in_model_area = False
    for row in ws.iter_rows(min_row=3, min_col=cols["multi"], max_col=cols["multi"] + 3, values_only=True):
        vals = [str(v).strip() if v is not None else "" for v in row]
        joined = "|".join(vals)
        if "车型" in joined and ("整车厂" in joined or "车企" in joined):
            in_model_area = True
            continue
        if not vals[0]:
            continue
        if in_model_area:
            if vals[1] and vals[2]:
                multi_model_oem[(vals[0], vals[1])] = vals[2]
                multi_model_oem[(strip_suffix(vals[0]), strip_suffix(vals[1]))] = vals[2]
        else:
            if vals[1]:
                multi_brand_oem[vals[0]] = vals[1]
                multi_brand_oem[strip_suffix(vals[0])] = vals[1]

    country_map = {}
    for row in ws.iter_rows(min_row=3, min_col=cols["country"], max_col=cols["country"] + 2, values_only=True):
        if row[0] and row[1] and row[2]:
            country_map[str(row[0]).strip()] = (str(row[1]).strip(), str(row[2]).strip())

    wb.close()
    return {
        "brand_series": brand_series,
        "brand_series_idx": build_idx(brand_series),
        "brand_oem": brand_oem,
        "brand_oem_idx": build_idx(brand_oem),
        "multi_brand_oem": multi_brand_oem,
        "multi_model_oem": multi_model_oem,
        "country_map": country_map,
    }


def clean_energy(value):
    s = str(value).strip() if value is not None else ""
    return "燃油车" if s == "传统能源" else s


def convert(main_path, raw_path):
    maps = load_maps(main_path)
    wb = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    ws.reset_dimensions()

    cleaned = []
    unmatched_country = set()
    unmatched_series = set()
    unmatched_oem = set()
    month_counter = Counter()
    qty_counter = Counter()

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    for row in rows:
        if not row or row[0] is None:
            continue
        year, month, country, brand, model, energy, qty = row[:7]
        try:
            year = int(year)
            month = int(month)
        except (TypeError, ValueError):
            continue
        country = str(country).strip()
        brand = str(brand).strip()
        model = str(model).strip() if model is not None else ""
        energy = clean_energy(energy)
        qty = int(qty or 0)

        geo = maps["country_map"].get(country)
        if not geo:
            unmatched_country.add(country)
            region, subregion = "未匹配", "未匹配"
        else:
            region, subregion = geo

        sk = maps["brand_series_idx"].get(brand) or maps["brand_series_idx"].get(strip_suffix(brand))
        series = maps["brand_series"].get(sk) if sk else "未匹配"
        if series == "未匹配":
            unmatched_series.add(brand)

        oem = (
            maps["multi_model_oem"].get((brand, model))
            or maps["multi_model_oem"].get((strip_suffix(brand), strip_suffix(model)))
            or maps["multi_brand_oem"].get(brand)
            or maps["multi_brand_oem"].get(strip_suffix(brand))
        )
        if not oem:
            ok = maps["brand_oem_idx"].get(brand) or maps["brand_oem_idx"].get(strip_suffix(brand))
            oem = maps["brand_oem"].get(ok) if ok else "未匹配"
        if oem == "未匹配":
            unmatched_oem.add(brand)

        cleaned.append([year, month, country, region, subregion, brand, model, energy, qty, series, oem])
        month_counter[f"{year}-{month:02d}"] += 1
        qty_counter[f"{year}-{month:02d}"] += qty

    wb.close()
    return {
        "header": header,
        "cleaned": cleaned,
        "cleaned_rows": len(cleaned),
        "month_rows": dict(sorted(month_counter.items())),
        "month_qty": {k: int(v) for k, v in sorted(qty_counter.items())},
        "unmatched_country": sorted(unmatched_country),
        "unmatched_series": sorted(unmatched_series),
        "unmatched_oem": sorted(unmatched_oem),
        "map_counts": {k: len(v) for k, v in maps.items()},
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--main", required=True)
    ap.add_argument("--raw", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()
    payload = convert(args.main, args.raw)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    public = {k: v for k, v in payload.items() if k != "cleaned"}
    print(json.dumps(public, ensure_ascii=False, indent=2))
    if payload["unmatched_country"] or payload["unmatched_series"] or payload["unmatched_oem"]:
        sys.exit(2)


if __name__ == "__main__":
    main()
