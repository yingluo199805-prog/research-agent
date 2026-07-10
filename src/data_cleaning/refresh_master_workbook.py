import argparse
import json
import os
import re
import shutil
import time
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook


BASE = Path.home() / "Desktop" / "收纳" / "08_代码与工具" / "投研报告_代码" / "海关出口看板"
MAIN = BASE / "海关出口数据-260612_working.xlsx"
DATASAVE = BASE / "data save"
MKLS_RAW = DATASAVE / "MarkLines_sales_data_cn.xlsx"
HG_RAW = DATASAVE / "2026_05各国出口量.xlsx"

LIGHT = "轻型车"
UNMATCHED = "未匹配"
UAE_SAUDI = {"阿联酋", "沙特阿拉伯"}
UAE_SAUDI_LEVELS = {"A", "B", "C", "D", "E", "F", "MPV", "Pickup Truck"}

MKLS_HEADER = [
    "年",
    "月",
    "国家/地区",
    "集团",
    "整车厂/品牌",
    "大洲",
    "区域",
    "标准车种",
    "标准动力类型",
    "车种",
    "级别",
    "车型",
    "新增区域划分",
    "标准车型",
    "系别",
    "车企划分",
    "数量",
]

HG_HEADER = ["年", "月", "国家", "区域", "细分区域", "品牌", "车型", "能源类型（分为）", "数量", "系别", "车企划分"]


def strip_suffix(name):
    return re.sub(r"\s*\(.*?\)\s*", "", str(name)).strip()


def fuzzy_index(mapping):
    idx = {}
    for key in mapping:
        idx[key] = key
        idx.setdefault(strip_suffix(key), key)
    return idx


def clean_text(value):
    return str(value).strip() if value is not None else ""


def as_int(value):
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def as_qty(value):
    if value in (None, "", "-"):
        return None
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    if v == 0:
        return None
    return int(v) if v.is_integer() else v


def ym_from_int(value):
    s = str(int(float(value)))
    return int(s[:4]), int(s[4:])


def is_uae_saudi_light(country, vehicle, level):
    if country not in UAE_SAUDI or vehicle != "N/A":
        return False
    level = str(level or "").strip()
    return level in UAE_SAUDI_LEVELS or level.startswith("SUV-")


def load_mkls_matching(main_path):
    wb = openpyxl.load_workbook(main_path, read_only=True, data_only=True)
    ws = wb["marklines匹配字段"]
    ws.reset_dimensions()
    vehicle_class = {}
    power_map = {}
    brand_series = {}
    brand_oem = {}
    country_geo = {}
    rebadge_rules = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) > 1 and row[0] and row[1]:
            vehicle_class[str(row[0]).strip()] = str(row[1]).strip()
        if len(row) > 4 and row[3] and row[4]:
            power_map[str(row[3]).strip()] = str(row[4]).strip()
        if len(row) > 10 and row[9] and row[10]:
            brand_series[str(row[9]).strip()] = str(row[10]).strip()
        if len(row) > 13 and row[12] and row[13]:
            brand_oem[str(row[12]).strip()] = str(row[13]).strip()
        if len(row) > 23 and row[20]:
            country_geo[str(row[20]).strip()] = {
                "continent": row[21],
                "region": row[22],
                "new_region": row[23],
            }
        if len(row) > 31 and row[25] and row[26] and row[30] and row[31]:
            key = (
                clean_text(row[25]),
                clean_text(row[26]),
                clean_text(row[27]),
                clean_text(row[28]),
            )
            rebadge_rules[key] = {
                "series": clean_text(row[30]),
                "oem": clean_text(row[31]),
            }
    wb.close()
    brand_series.setdefault("吉利银河 (Geely Galaxy)", "中系")
    brand_oem.setdefault("吉利银河 (Geely Galaxy)", "吉利汽车 H0175")
    country_geo.setdefault("阿联酋", {"continent": "亚洲", "region": "西亚", "new_region": "中东"})
    country_geo.setdefault("沙特阿拉伯", {"continent": "亚洲", "region": "西亚", "new_region": "中东"})
    return {
        "vehicle_class": vehicle_class,
        "power_map": power_map,
        "brand_series": brand_series,
        "brand_series_idx": fuzzy_index(brand_series),
        "brand_oem": brand_oem,
        "brand_oem_idx": fuzzy_index(brand_oem),
        "country_geo": country_geo,
        "rebadge_rules": rebadge_rules,
    }


def lookup_mkls_rebadge(matching, country, brand, model):
    model = clean_text(model)
    candidates = []
    if model:
        candidates.extend(
            [
                (country, brand, model, model),
                (country, brand, model, ""),
                (country, brand, "", model),
            ]
        )
    else:
        candidates.append((country, brand, "", ""))
    for key in candidates:
        rule = matching["rebadge_rules"].get(key)
        if rule:
            return rule
    return None


def build_mkls_temp(main_path, raw_path, out_path):
    t0 = time.time()
    matching = load_mkls_matching(main_path)
    wb_raw = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)
    ws_raw = wb_raw[wb_raw.sheetnames[0]]
    ws_raw.reset_dimensions()
    header = next(ws_raw.iter_rows(min_row=2, max_row=2, values_only=True))
    month_cols = []
    month_sums = Counter()
    for idx, value in enumerate(header):
        month = as_int(value)
        if idx >= 7 and month and month >= 202401:
            month_cols.append((idx, month))

    for row in ws_raw.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        for idx, month in month_cols:
            if idx < len(row):
                qty = as_qty(row[idx])
                if qty:
                    month_sums[month] += qty
    active_months = sorted(month for month, total in month_sums.items() if total)
    active_ym = {ym_from_int(month) for month in active_months}
    wb_raw.close()

    stats = {
        "active_months": active_months,
        "active_ym": sorted(active_ym),
        "kept_old_rows_before_2024": 0,
        "skipped_old_rows_2024_plus": 0,
        "new_rows": 0,
        "new_by_ym": {},
        "qty_sum_by_month": {str(m): 0 for m in active_months},
        "unmatched_country": [],
        "unmatched_series": [],
        "unmatched_oem": [],
        "uae_saudi_included_rows": 0,
        "uae_saudi_included_qty": {},
        "uae_saudi_excluded_unclassified_rows": 0,
        "uae_saudi_excluded_unclassified_qty": {},
        "rebadge_override_rows": 0,
        "rebadge_override_qty": {},
        "elapsed_build_sec": None,
    }

    out_wb = Workbook(write_only=True)
    out_ws = out_wb.create_sheet("mkls-CV")
    out_ws.append(MKLS_HEADER)

    wb_old = openpyxl.load_workbook(main_path, read_only=True, data_only=True)
    ws_old = wb_old["mkls-CV"]
    ws_old.reset_dimensions()
    for row in ws_old.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        year = as_int(row[0])
        if year is None:
            continue
        if year >= 2024:
            stats["skipped_old_rows_2024_plus"] += 1
            continue
        out_ws.append(list(row[:17]))
        stats["kept_old_rows_before_2024"] += 1
    wb_old.close()

    unmatched_country = set()
    unmatched_series = set()
    unmatched_oem = set()
    new_by_ym = Counter()
    qty_sum_by_month = Counter()
    included_qty = Counter()
    excluded_qty = Counter()
    rebadge_qty = Counter()

    wb_raw = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)
    ws_raw = wb_raw[wb_raw.sheetnames[0]]
    ws_raw.reset_dimensions()
    raw_header = next(ws_raw.iter_rows(min_row=2, max_row=2, values_only=True))
    target_cols = [(idx, as_int(value)) for idx, value in enumerate(raw_header) if idx >= 7 and as_int(value) in active_months]

    for row in ws_raw.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 7 or row[0] is None:
            continue
        country = str(row[0]).strip()
        if country == "中国":
            continue
        group = str(row[1]).strip() if row[1] else ""
        brand = str(row[2]).strip() if row[2] else ""
        raw_vehicle = str(row[3]).strip() if row[3] else ""
        level = row[4]
        model = row[5]
        raw_power = str(row[6]).strip() if row[6] else ""
        standard_vehicle = matching["vehicle_class"].get(raw_vehicle)
        special_included = False
        if standard_vehicle != LIGHT:
            if is_uae_saudi_light(country, raw_vehicle, level):
                standard_vehicle = LIGHT
                special_included = True
            elif country in UAE_SAUDI and raw_vehicle == "N/A" and str(level or "").strip() == "不可分级":
                row_total = 0
                for idx, month in target_cols:
                    if idx < len(row):
                        qty = as_qty(row[idx])
                        if qty:
                            row_total += qty
                if row_total:
                    stats["uae_saudi_excluded_unclassified_rows"] += 1
                    excluded_qty[country] += row_total
                continue
            else:
                continue

        standard_power = matching["power_map"].get(raw_power, "ICE")
        geo = matching["country_geo"].get(country)
        if geo is None:
            unmatched_country.add(country)
            geo = {"continent": None, "region": None, "new_region": None}
        series_key = matching["brand_series_idx"].get(brand) or matching["brand_series_idx"].get(strip_suffix(brand))
        series = matching["brand_series"].get(series_key) if series_key else None
        if series is None and brand:
            unmatched_series.add(brand)
        oem_key = matching["brand_oem_idx"].get(brand) or matching["brand_oem_idx"].get(strip_suffix(brand))
        oem = matching["brand_oem"].get(oem_key) if oem_key else None
        if oem is None and brand:
            unmatched_oem.add(brand)
        rebadge = lookup_mkls_rebadge(matching, country, brand, model)
        if rebadge:
            series = rebadge["series"]
            oem = rebadge["oem"]

        for idx, month in target_cols:
            if idx >= len(row):
                continue
            qty = as_qty(row[idx])
            if qty is None:
                continue
            year, month_num = ym_from_int(month)
            out_ws.append(
                [
                    year,
                    month_num,
                    country,
                    group,
                    brand,
                    geo["continent"],
                    geo["region"],
                    standard_vehicle,
                    standard_power,
                    raw_vehicle,
                    level,
                    model,
                    geo["new_region"],
                    model,
                    series,
                    oem,
                    qty,
                ]
            )
            stats["new_rows"] += 1
            new_by_ym[(year, month_num)] += 1
            qty_sum_by_month[month] += qty
            if special_included:
                stats["uae_saudi_included_rows"] += 1
                included_qty[country] += qty
            if rebadge:
                stats["rebadge_override_rows"] += 1
                rebadge_qty[oem] += qty
    wb_raw.close()

    out_wb.save(out_path)
    stats["new_by_ym"] = {f"{y}-{m:02d}": c for (y, m), c in sorted(new_by_ym.items())}
    stats["qty_sum_by_month"] = {str(k): int(v) if float(v).is_integer() else v for k, v in sorted(qty_sum_by_month.items())}
    stats["unmatched_country"] = sorted(unmatched_country)
    stats["unmatched_series"] = sorted(unmatched_series)
    stats["unmatched_oem"] = sorted(unmatched_oem)
    stats["uae_saudi_included_qty"] = {k: int(v) if float(v).is_integer() else v for k, v in sorted(included_qty.items())}
    stats["uae_saudi_excluded_unclassified_qty"] = {k: int(v) if float(v).is_integer() else v for k, v in sorted(excluded_qty.items())}
    stats["rebadge_override_qty"] = {k: int(v) if float(v).is_integer() else v for k, v in sorted(rebadge_qty.items())}
    stats["total_rows_after"] = stats["kept_old_rows_before_2024"] + stats["new_rows"]
    stats["elapsed_build_sec"] = round(time.time() - t0, 1)
    return stats


def load_hg_maps(main_path):
    wb = openpyxl.load_workbook(main_path, read_only=True, data_only=True)
    ws = wb["海关匹配字段"]
    ws.reset_dimensions()
    brand_series = {}
    brand_oem = {}
    multi_brands = set()
    multi_model_oem = {}
    country_map = {}
    in_model = False
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) > 1 and row[0] and row[1]:
            brand_series[str(row[0]).strip()] = str(row[1]).strip()
        if len(row) > 4 and row[3] and row[4]:
            brand_oem[str(row[3]).strip()] = str(row[4]).strip()
        vals = [str(v).strip() if v is not None else "" for v in row[6:10]]
        joined = "|".join(vals)
        if "车型" in joined and ("映射" in joined or "整车厂" in joined):
            in_model = True
        elif vals[0]:
            if in_model:
                if vals[1] and vals[2]:
                    multi_model_oem[(vals[0], vals[1])] = vals[2]
                    multi_model_oem[(strip_suffix(vals[0]), strip_suffix(vals[1]))] = vals[2]
            elif vals[1]:
                multi_brands.add(vals[0])
                multi_brands.add(strip_suffix(vals[0]))
        if len(row) > 12 and row[10] and row[11] and row[12]:
            country_map[str(row[10]).strip()] = (str(row[11]).strip(), str(row[12]).strip())
    wb.close()
    return {
        "brand_series": brand_series,
        "brand_series_idx": fuzzy_index(brand_series),
        "brand_oem": brand_oem,
        "brand_oem_idx": fuzzy_index(brand_oem),
        "multi_brands": multi_brands,
        "multi_model_oem": multi_model_oem,
        "country_map": country_map,
    }


def clean_hg_energy(value):
    s = str(value).strip() if value is not None else ""
    return "燃油车" if s == "传统能源" else s


def build_hg_temp(main_path, raw_path, out_path):
    maps = load_hg_maps(main_path)
    raw_wb = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)
    raw_ws = raw_wb[raw_wb.sheetnames[0]]
    raw_ws.reset_dimensions()
    cleaned = []
    skipped_domestic_rows = 0
    skipped_domestic_qty = 0
    unmatched_country = set()
    unmatched_series = set()
    unmatched_oem = set()
    month_counter = Counter()
    qty_counter = Counter()
    for row in raw_ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        year, month, country, brand, model, energy, qty = row[:7]
        year = as_int(year)
        month = as_int(month)
        if not year or not month:
            continue
        country = str(country).strip()
        if country == "中国":
            skipped_domestic_rows += 1
            skipped_domestic_qty += int(float(qty or 0))
            continue
        brand = str(brand).strip()
        model = str(model).strip() if model is not None else ""
        qty = int(float(qty or 0))
        geo = maps["country_map"].get(country)
        if not geo:
            region, subregion = UNMATCHED, UNMATCHED
            unmatched_country.add(country)
        else:
            region, subregion = geo
        sk = maps["brand_series_idx"].get(brand) or maps["brand_series_idx"].get(strip_suffix(brand))
        series = maps["brand_series"].get(sk) if sk else UNMATCHED
        if series == UNMATCHED:
            unmatched_series.add(brand)
        if brand in maps["multi_brands"] or strip_suffix(brand) in maps["multi_brands"]:
            oem = maps["multi_model_oem"].get((brand, model)) or maps["multi_model_oem"].get((strip_suffix(brand), strip_suffix(model)))
            if not oem:
                oem = UNMATCHED
        else:
            ok = maps["brand_oem_idx"].get(brand) or maps["brand_oem_idx"].get(strip_suffix(brand))
            oem = maps["brand_oem"].get(ok) if ok else UNMATCHED
        if oem == UNMATCHED:
            unmatched_oem.add(f"{brand}|{model}")
        cleaned.append([year, month, country, region, subregion, brand, model, clean_hg_energy(energy), qty, series, oem])
        month_counter[f"{year}-{month:02d}"] += 1
        qty_counter[f"{year}-{month:02d}"] += qty
    raw_wb.close()

    out_wb = Workbook(write_only=True)
    out_ws = out_wb.create_sheet("海关CV")
    out_ws.append(HG_HEADER)
    kept_old = 0
    removed_existing = 0
    wb_old = openpyxl.load_workbook(main_path, read_only=True, data_only=True)
    ws_old = wb_old["海关CV"]
    ws_old.reset_dimensions()
    target_ym = {(r[0], r[1]) for r in cleaned}
    for row in ws_old.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        ym = (as_int(row[0]), as_int(row[1]))
        if ym in target_ym:
            removed_existing += 1
            continue
        out_ws.append(list(row[:11]))
        kept_old += 1
    wb_old.close()
    for row in cleaned:
        out_ws.append(row)
    out_wb.save(out_path)

    return {
        "cleaned_rows": len(cleaned),
        "month_rows": dict(sorted(month_counter.items())),
        "month_qty": {k: int(v) for k, v in sorted(qty_counter.items())},
        "kept_old_rows": kept_old,
        "removed_existing_target_month_rows": removed_existing,
        "total_rows_after": kept_old + len(cleaned),
        "skipped_domestic_rows": skipped_domestic_rows,
        "skipped_domestic_qty": skipped_domestic_qty,
        "unmatched_country": sorted(unmatched_country),
        "unmatched_series": sorted(unmatched_series),
        "unmatched_oem": sorted(unmatched_oem),
    }


def workbook_sheet_path(xlsx_path, sheet_name):
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        workbook = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid = None
    for sheet in workbook.find("main:sheets", ns):
        if sheet.attrib.get("name") == sheet_name:
            rid = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            break
    if rid is None:
        raise RuntimeError(f"sheet not found: {sheet_name}")
    for rel in rels:
        if rel.attrib.get("Id") == rid:
            target = rel.attrib["Target"].lstrip("/")
            return target if target.startswith("xl/") else "xl/" + target
    raise RuntimeError(f"sheet rel not found: {sheet_name}")


def xml_escape(value):
    return str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def col_letter(col_idx):
    s = ""
    n = col_idx + 1
    while n:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s


def encode_row(row_num, values, col_count):
    parts = [f'<row r="{row_num}">']
    for col_idx, value in enumerate(values[:col_count]):
        if value is None:
            continue
        ref = col_letter(col_idx) + str(row_num)
        if isinstance(value, bool):
            parts.append(f'<c r="{ref}" t="b"><v>{1 if value else 0}</v></c>')
        elif isinstance(value, (int, float)):
            if isinstance(value, float) and value.is_integer():
                value = int(value)
            parts.append(f'<c r="{ref}"><v>{value}</v></c>')
        else:
            parts.append(f'<c r="{ref}" t="inlineStr"><is><t xml:space="preserve">{xml_escape(value)}</t></is></c>')
    parts.append("</row>")
    return "".join(parts).encode("utf-8")


def sheet_xml_from_temp(temp_path, sheet_name, col_count, last_col_letter):
    wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    ws.reset_dimensions()
    total_rows = sum(1 for _ in ws.iter_rows(values_only=True))
    wb.close()
    opening = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
        'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac" '
        'mc:Ignorable="x14ac">'
        f'<dimension ref="A1:{last_col_letter}{total_rows}"/>'
        '<sheetViews><sheetView workbookViewId="0"/></sheetViews>'
        '<sheetFormatPr defaultRowHeight="15" x14ac:dyDescent="0.25"/>'
        "<sheetData>"
    ).encode("utf-8")
    closing = b"</sheetData></worksheet>"
    return opening, closing, total_rows


def replace_sheets(main_path, replacements):
    backup = main_path.with_name(main_path.stem + f"_preDatasave260702Backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    shutil.copy2(main_path, backup)
    tmp = main_path.with_suffix(".xlsx.datasave260702.tmp")
    sheet_payload = {}
    for sheet_name, spec in replacements.items():
        sheet_path = workbook_sheet_path(main_path, sheet_name)
        opening, closing, total_rows = sheet_xml_from_temp(spec["temp"], sheet_name, spec["col_count"], spec["last_col"])
        sheet_payload[sheet_path] = {**spec, "opening": opening, "closing": closing, "total_rows": total_rows}

    with zipfile.ZipFile(main_path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED, compresslevel=6, allowZip64=True) as zout:
        for item in zin.infolist():
            payload = sheet_payload.get(item.filename)
            if payload:
                with zout.open(item, "w", force_zip64=True) as fp:
                    fp.write(payload["opening"])
                    wb = openpyxl.load_workbook(payload["temp"], read_only=True, data_only=True)
                    ws = wb[payload["sheet"]]
                    ws.reset_dimensions()
                    row_num = 1
                    for row in ws.iter_rows(values_only=True):
                        if not row or all(v is None for v in row[: payload["col_count"]]):
                            continue
                        fp.write(encode_row(row_num, row, payload["col_count"]))
                        row_num += 1
                    wb.close()
                    fp.write(payload["closing"])
            else:
                zout.writestr(item, zin.read(item.filename))
    os.replace(tmp, main_path)
    return {"backup": str(backup), "sheet_rows": {k: v["total_rows"] for k, v in sheet_payload.items()}}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--main", required=True, type=Path)
    parser.add_argument("--mkls-raw", required=True, type=Path)
    parser.add_argument("--hg-raw", required=True, type=Path)
    parser.add_argument("--writeback", action="store_true")
    parser.add_argument("--report", default="datasave_update_report.json")
    args = parser.parse_args()
    report_path = Path(args.report).resolve()
    report_path.parent.mkdir(parents=True, exist_ok=True)
    main_path = args.main.resolve()
    mkls_raw = args.mkls_raw.resolve()
    hg_raw = args.hg_raw.resolve()
    for path in (main_path, mkls_raw, hg_raw):
        if not path.is_file():
            raise FileNotFoundError(path)
    temp_mkls = report_path.with_name("datasave_update_mkls_temp.xlsx")
    temp_hg = report_path.with_name("datasave_update_customs_temp.xlsx")

    report = {
        "main": str(main_path),
        "mkls_raw": str(mkls_raw),
        "hg_raw": str(hg_raw),
        "mkls": build_mkls_temp(main_path, mkls_raw, temp_mkls),
        "haiguan": build_hg_temp(main_path, hg_raw, temp_hg),
        "status": "validated",
    }
    blocked = []
    for section in ("mkls", "haiguan"):
        for key in ("unmatched_country", "unmatched_series", "unmatched_oem"):
            if report[section].get(key):
                blocked.append(f"{section}.{key}")
    if blocked:
        report["status"] = "blocked_unmatched"
        report["blocked"] = blocked
    elif args.writeback:
        report["writeback"] = replace_sheets(
            main_path,
            {
                "mkls-CV": {"temp": temp_mkls, "sheet": "mkls-CV", "col_count": 17, "last_col": "Q"},
                "海关CV": {"temp": temp_hg, "sheet": "海关CV", "col_count": 11, "last_col": "K"},
            },
        )
        report["status"] = "written"

    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    public = json.loads(json.dumps(report, ensure_ascii=False))
    print(json.dumps(public, ensure_ascii=False, indent=2))
    if report["status"] == "blocked_unmatched":
        raise SystemExit(2)


if __name__ == "__main__":
    main()
