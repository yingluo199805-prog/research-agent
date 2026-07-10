# -*- coding: utf-8 -*-
"""Fix M1-5 update: insert 4 cols after 2026Q1 (user column order), fill MKLS data."""

from __future__ import annotations

import argparse
import json
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import pythoncom
import win32com.client as win32

from m15_layout import M15_WIDTH, insert_col_for_q, m15_col_after_inserts, q1_col_after_inserts
from update_workbook_m15_inline_fast import (
    SHEET_TO_REGION,
    build_change_plan,
    build_table_plan,
    cell_addr,
    col_letter,
    text,
)


XL_SHIFT_TO_RIGHT = -4161
XL_CALC_AUTOMATIC = -4105


@dataclass
class FillPlan:
    start_row: int
    end_row: int
    m15_col: int
    matrix: list[list[Any]]
    is_share: bool


def collect_qcols(path: Path, sheet_name: str) -> list[int]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    qcols = sorted(
        {
            cell.column
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row)
            for cell in row
            if cell.value == "2026Q1"
        }
    )
    wb.close()
    return qcols


def rebuild_matrix_formulas(matrix: list[list[Any]], start_row: int, m15_col: int, share: bool) -> list[list[Any]]:
    col_25 = m15_col
    col_26 = m15_col + 1
    out: list[list[Any]] = []
    for i, row_vals in enumerate(matrix):
        r = start_row + i
        if i == 0:
            out.append(list(row_vals))
            continue
        rebuilt = list(row_vals[:2])
        rebuilt.append(f"={cell_addr(r, col_26)}-{cell_addr(r, col_25)}")
        if share:
            rebuilt.append("" if len(row_vals) < 4 else row_vals[3])
        else:
            rebuilt.append(f'=IFERROR({cell_addr(r, col_26)}/{cell_addr(r, col_25)}-1,"")')
        out.append(rebuilt)
    return out


def build_fill_plans(
    cache: dict[str, Any],
    target_path: Path,
    sheet_name: str,
    region: str,
    orig_qcols: list[int],
) -> list[FillPlan]:
    wb = openpyxl.load_workbook(target_path, read_only=False, data_only=False)
    ws = wb[sheet_name]
    fills: list[FillPlan] = []

    for orig_qcol in orig_qcols:
        m15_col = m15_col_after_inserts(orig_qcol, orig_qcols)
        for header_row in range(1, ws.max_row + 1):
            if ws.cell(header_row, orig_qcol).value != "2026Q1":
                continue
            plan = build_table_plan(ws, cache, sheet_name, region, header_row, orig_qcol)
            if plan:
                fills.append(
                    FillPlan(
                        start_row=plan.start_row,
                        end_row=plan.end_row,
                        m15_col=m15_col,
                        matrix=rebuild_matrix_formulas(plan.matrix, plan.start_row, m15_col, plan.is_share),
                        is_share=plan.is_share,
                    )
                )

    for header_row in range(1, ws.max_row + 1):
        for orig_qcol in range(1, ws.max_column + 1):
            value = text(ws.cell(header_row, orig_qcol).value)
            if "2026Q1相较于2025Q1" not in value:
                continue
            plan = build_change_plan(ws, cache, sheet_name, region, header_row, orig_qcol)
            if not plan:
                continue
            m15_col = m15_col_after_inserts(orig_qcol, orig_qcols)
            fills.append(
                FillPlan(
                    start_row=plan.start_row,
                    end_row=plan.end_row,
                    m15_col=m15_col,
                    matrix=rebuild_matrix_formulas(plan.matrix, plan.start_row, m15_col, True),
                    is_share=True,
                )
            )

    wb.close()
    return fills


def insert_full_columns(ws: Any, orig_qcols: list[int]) -> None:
    for orig_q in sorted(orig_qcols, reverse=True):
        ic = insert_col_for_q(orig_q)
        ws.Range(f"{col_letter(ic)}:{col_letter(ic + M15_WIDTH - 1)}").EntireColumn.Insert(Shift=XL_SHIFT_TO_RIGHT)


def write_matrix(ws: Any, row: int, col: int, matrix: list[list[Any]]) -> None:
    nrows = len(matrix)
    ncols = len(matrix[0])
    address = f"{col_letter(col)}{row}:{col_letter(col + ncols - 1)}{row + nrows - 1}"
    ws.Range(address).Formula = tuple(tuple(r) for r in matrix)


def apply_table_format(ws: Any, fill: FillPlan) -> None:
    qcol = fill.m15_col - 1
    m15_col = fill.m15_col
    ncols = len(fill.matrix[0])
    header_src = ws.Range(f"{col_letter(qcol)}{fill.start_row}")
    header_src.Copy()
    ws.Range(
        f"{col_letter(m15_col)}{fill.start_row}:"
        f"{col_letter(m15_col + ncols - 1)}{fill.start_row}"
    ).PasteSpecial(Paste=-4122)
    if fill.end_row > fill.start_row:
        if fill.is_share:
            ws.Range(
                f"{col_letter(m15_col)}{fill.start_row + 1}:"
                f"{col_letter(m15_col + ncols - 1)}{fill.end_row}"
            ).NumberFormat = "0.0%"
        else:
            ws.Range(
                f"{col_letter(m15_col)}{fill.start_row + 1}:"
                f"{col_letter(m15_col + 2)}{fill.end_row}"
            ).NumberFormat = "#,##0"
            ws.Range(
                f"{col_letter(m15_col + 3)}{fill.start_row + 1}:"
                f"{col_letter(m15_col + 3)}{fill.end_row}"
            ).NumberFormat = "0.0%"
    for offset, width in enumerate((11, 11, 12, 12)):
        if offset < ncols:
            ws.Columns(m15_col + offset).ColumnWidth = width


def sheet_has_m15(ws: Any) -> bool:
    used = ws.UsedRange
    found = used.Find("2025M1-5")
    return found is not None


def update_sheet(
    cache: dict[str, Any],
    plan_source: Path,
    output_path: Path,
    sheet_name: str,
    *,
    in_place: bool = False,
    skip_if_m15: bool = False,
) -> bool:
    region = SHEET_TO_REGION[sheet_name]
    orig_qcols = collect_qcols(plan_source, sheet_name)
    if not orig_qcols:
        return False

    if not in_place:
        if output_path.exists():
            output_path.unlink()
        shutil.copy2(plan_source, output_path)

    fills = build_fill_plans(cache, plan_source, sheet_name, region, orig_qcols)

    pythoncom.CoInitialize()
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    try:
        wb = excel.Workbooks.Open(str(output_path.resolve()))
        ws = wb.Worksheets(sheet_name)
        if skip_if_m15 and sheet_has_m15(ws):
            wb.Close(SaveChanges=False)
            return False
        insert_full_columns(ws, orig_qcols)
        for fill in fills:
            write_matrix(ws, fill.start_row, fill.m15_col, fill.matrix)
            apply_table_format(ws, fill)
        excel.CutCopyMode = False
        excel.Calculation = XL_CALC_AUTOMATIC
        wb.Save()
        wb.Close(SaveChanges=True)
    finally:
        excel.ScreenUpdating = True
        excel.DisplayAlerts = True
        excel.Quit()
        pythoncom.CoUninitialize()
    return True


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--cache", required=True, type=Path)
    parser.add_argument("--target-workbook", required=True, type=Path)
    parser.add_argument("--output-workbook", required=True, type=Path)
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--skip-if-m15", action="store_true")
    args = parser.parse_args()
    cache = json.loads(args.cache.read_text(encoding="utf-8"))
    update_sheet(cache, args.target_workbook, args.output_workbook, args.sheet, args.skip_if_m15)
    print(f"Updated sheet: {args.sheet}")


if __name__ == "__main__":
    main()
