# -*- coding: utf-8 -*-
"""Build 2025M1-5 / 2026M1-5 MKLS aggregates for the regional workbook."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


YEARS = (2025, 2026)
MONTHS = {1, 2, 3, 4, 5}
POWERS = ("ICE", "HV", "PHV", "EV")
SERIES = ("欧系", "日系", "韩系", "美系", "中系", "其他")

REGION_SHEETS = {
    "欧洲": {
        "continent": "欧洲",
        "overall_sheet": "欧洲总体分析",
        "country_sheet": "欧洲分国家",
        "core_sheet": "欧洲-核心数据",
        "exclude_countries": {"俄罗斯"},
    },
    "亚洲": {
        "continent": "亚洲",
        "overall_sheet": "亚洲总体分析",
        "country_sheet": "亚洲分国家分析",
        "core_sheet": "亚洲-核心数据",
        "exclude_countries": {"中国"},
    },
    "北美洲": {
        "continent": "北美洲",
        "overall_sheet": "北美洲总体分析",
        "country_sheet": "北美洲分国家",
        "core_sheet": "北美洲-核心数据",
        "exclude_countries": set(),
    },
    "南美洲": {
        "continent": "南美洲",
        "overall_sheet": "南美洲总体分析",
        "country_sheet": "南美洲分国家",
        "core_sheet": "南美洲-核心数据",
        "exclude_countries": set(),
    },
    "大洋洲": {
        "continent": "大洋洲",
        "overall_sheet": "大洋洲总体分析",
        "country_sheet": "大洋洲分国家",
        "core_sheet": "大洋洲-核心数据",
        "exclude_countries": set(),
    },
    "非洲": {
        "continent": "非洲",
        "overall_sheet": "非洲总体分析",
        "country_sheet": "非洲分国家",
        "core_sheet": "非洲-核心数据",
        "exclude_countries": set(),
    },
    "俄罗斯": {
        "continent": "欧洲",
        "overall_sheet": None,
        "country_sheet": None,
        "core_sheet": "俄罗斯-核心数据",
        "countries": ["俄罗斯"],
        "exclude_countries": set(),
    },
}

COUNTRY_ALIASES = {
    "台湾": "中国台湾",
}


def ydict() -> dict[str, float]:
    return {"2025": 0.0, "2026": 0.0}


def nested_year_dict() -> defaultdict[str, dict[str, float]]:
    return defaultdict(ydict)


def add_year(bucket: dict[str, float], year: int, qty: float) -> None:
    if year in YEARS:
        bucket[str(year)] = bucket.get(str(year), 0.0) + qty


def normalize_country(country: Any) -> str:
    value = "" if country is None else str(country).strip()
    return COUNTRY_ALIASES.get(value, value)


def clean_dim(value: Any, fallback: str = "其他") -> str:
    if value is None:
        return fallback
    text = str(value).strip()
    if not text or text.upper() in {"N/A", "NA", "NONE"}:
        return fallback
    return text


def read_overall_countries(wb: openpyxl.Workbook, sheet_name: str | None) -> list[str]:
    if not sheet_name or sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    countries: list[str] = []
    for row in range(5, min(ws.max_row, 120) + 1):
        value = ws.cell(row=row, column=1).value
        if value is None:
            continue
        text = str(value).strip()
        if "合计" in text:
            break
        if any(mark in text for mark in ("占比", "销量", "动力类型", "系别", "市占率")):
            continue
        countries.append(text)
    return countries


def read_core_countries(wb: openpyxl.Workbook, sheet_name: str | None) -> list[str]:
    if not sheet_name or sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    countries: list[str] = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value is None:
            continue
        text = str(value).strip()
        if "销量分析" in text:
            countries.append(text.split("销量分析", 1)[0])
    return countries


def build_region_config(target_workbook: Path) -> dict[str, dict[str, Any]]:
    wb = openpyxl.load_workbook(target_workbook, read_only=False, data_only=False)
    configs: dict[str, dict[str, Any]] = {}
    for region, raw in REGION_SHEETS.items():
        cfg = dict(raw)
        if "countries" not in cfg:
            cfg["countries"] = read_overall_countries(wb, cfg.get("overall_sheet"))
        cfg["core_countries"] = read_core_countries(wb, cfg.get("country_sheet"))
        cfg["countries"] = [normalize_country(x) for x in cfg.get("countries", [])]
        cfg["core_countries"] = [normalize_country(x) for x in cfg.get("core_countries", [])]
        configs[region] = cfg
    wb.close()
    return configs


def new_region_bucket() -> dict[str, Any]:
    return {
        "total": ydict(),
        "countries": defaultdict(ydict),
        "power": nested_year_dict(),
        "series": nested_year_dict(),
        "power_series": defaultdict(nested_year_dict),
        "company": nested_year_dict(),
        "models_by_power": defaultdict(nested_year_dict),
        "country_details": defaultdict(
            lambda: {
                "total": ydict(),
                "power": nested_year_dict(),
                "series": nested_year_dict(),
                "power_series": defaultdict(nested_year_dict),
                "company": nested_year_dict(),
                "models_by_power": defaultdict(nested_year_dict),
            }
        ),
    }


def regularize(obj: Any) -> Any:
    if isinstance(obj, defaultdict):
        return {k: regularize(v) for k, v in obj.items()}
    if isinstance(obj, dict):
        return {k: regularize(v) for k, v in obj.items()}
    if isinstance(obj, set):
        return sorted(obj)
    return obj


def build_cache(source_mkls: Path, target_workbook: Path, output: Path) -> None:
    configs = build_region_config(target_workbook)
    country_to_region: dict[str, str] = {}
    for region, cfg in configs.items():
        for country in cfg["countries"]:
            country_to_region[country] = region

    cache: dict[str, Any] = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_mkls": str(source_mkls),
        "target_workbook": str(target_workbook),
        "years": list(YEARS),
        "months": sorted(MONTHS),
        "regions_config": configs,
        "regions": defaultdict(new_region_bucket),
        "unmatched_config_countries": {},
        "seen_countries": defaultdict(ydict),
        "row_count": 0,
        "used_row_count": 0,
    }

    wb = openpyxl.load_workbook(source_mkls, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Column positions in mkls-CV.
    Y, M, COUNTRY, GROUP, BRAND, CONT, REGION, STD_CAR, STD_POWER, CAR, LEVEL, MODEL, NEW_REGION, STD_MODEL, SERIES_COL, COMPANY, QTY = range(17)

    next(ws.iter_rows(min_row=1, max_row=1))
    for row in ws.iter_rows(min_row=2, values_only=True):
        cache["row_count"] += 1
        try:
            year = int(row[Y])
            month = int(row[M])
        except (TypeError, ValueError):
            continue
        if year not in YEARS or month not in MONTHS:
            continue

        country = normalize_country(row[COUNTRY])
        region = country_to_region.get(country)
        if not region:
            continue

        qty_value = row[QTY]
        if qty_value is None:
            continue
        try:
            qty = float(qty_value)
        except (TypeError, ValueError):
            continue
        if not qty:
            continue

        cache["used_row_count"] += 1
        power = clean_dim(row[STD_POWER], "其他")
        series = clean_dim(row[SERIES_COL], "其他")
        if series not in SERIES:
            series = "其他"
        company = clean_dim(row[COMPANY], "其他")
        model = clean_dim(row[STD_MODEL] or row[MODEL], "")
        model_key = f"{company}||{model}" if model else ""

        region_bucket = cache["regions"][region]
        add_year(region_bucket["total"], year, qty)
        add_year(region_bucket["countries"][country], year, qty)
        add_year(region_bucket["power"][power], year, qty)
        add_year(region_bucket["series"][series], year, qty)
        add_year(region_bucket["power_series"][power][series], year, qty)
        add_year(region_bucket["company"][company], year, qty)
        if model_key:
            add_year(region_bucket["models_by_power"][power][model_key], year, qty)

        detail = region_bucket["country_details"][country]
        add_year(detail["total"], year, qty)
        add_year(detail["power"][power], year, qty)
        add_year(detail["series"][series], year, qty)
        add_year(detail["power_series"][power][series], year, qty)
        add_year(detail["company"][company], year, qty)
        if model_key:
            add_year(detail["models_by_power"][power][model_key], year, qty)

        add_year(cache["seen_countries"][country], year, qty)

    wb.close()

    for region, cfg in configs.items():
        missing = []
        for country in cfg["countries"]:
            values = cache["regions"][region]["countries"].get(country, ydict())
            if not values.get("2025") and not values.get("2026"):
                missing.append(country)
        cache["unmatched_config_countries"][region] = missing

    output.write_text(json.dumps(regularize(cache), ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-mkls", required=True, type=Path)
    parser.add_argument("--target-workbook", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    build_cache(args.source_mkls, args.target_workbook, args.output)


if __name__ == "__main__":
    main()
