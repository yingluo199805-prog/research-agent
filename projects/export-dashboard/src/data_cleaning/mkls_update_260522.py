import argparse
import json
import os
import re
import shutil
import sys
import time
import zipfile
from collections import Counter
from datetime import datetime

import openpyxl
from openpyxl import Workbook


HEADER17 = [
    "年",
    "月",
    "国家/地区",
    "集团",
    "整车厂/品牌",
    "大洲",
    "区域",
    "标准车种",
    "标准动力类型",
    "车种",
    "级别",
    "车型",
    "新增区域划分",
    "标准车型",
    "系别",
    "车企划分",
    "数量",
]


def strip_suffix(name):
    return re.sub(r"\s*\(.*?\)\s*", "", str(name)).strip()


def ym_from_int(value):
    s = str(int(float(value)))
    if not re.fullmatch(r"\d{6}", s):
        raise ValueError(value)
    return int(s[:4]), int(s[4:])


def load_matching(main_path):
    wb = openpyxl.load_workbook(main_path, read_only=True, data_only=True)
    ws = wb["marklines匹配字段"]
    ws.reset_dimensions()
    rows = [r for r in ws.iter_rows(values_only=True)]
    wb.close()

    vehicle_class = {}
    power_map = {}
    brand_series = {}
    brand_oem = {}
    country_geo = {}

    for r in rows[2:]:
        if len(r) > 1 and r[0] and r[1]:
            vehicle_class[str(r[0]).strip()] = str(r[1]).strip()
        if len(r) > 4 and r[3] and r[4]:
            power_map[str(r[3]).strip()] = str(r[4]).strip()
        if len(r) > 10 and r[9] and r[10]:
            brand_series[str(r[9]).strip()] = str(r[10]).strip()
        if len(r) > 13 and r[12] and r[13]:
            brand_oem[str(r[12]).strip()] = str(r[13]).strip()
        if len(r) > 23 and r[20]:
            country_geo[str(r[20]).strip()] = {
                "continent": r[21],
                "region": r[22],
                "new_region": r[23],
            }

    return {
        "vehicle_class": vehicle_class,
        "power_map": power_map,
        "brand_series": brand_series,
        "brand_series_idx": fuzzy_index(brand_series),
        "brand_oem": brand_oem,
        "brand_oem_idx": fuzzy_index(brand_oem),
        "country_geo": country_geo,
    }


def fuzzy_index(mapping):
    idx = {}
    for key in mapping:
        idx[key] = key
        idx.setdefault(strip_suffix(key), key)
    return idx


def build_merged(main_path, raw_path, out_path, target_months):
    target_set = {int(x) for x in target_months}
    target_ym = {ym_from_int(x) for x in target_set}
    matching = load_matching(main_path)
    stats = {
        "target_months": sorted(target_set),
        "target_ym": sorted(target_ym),
        "kept_rows": 0,
        "skipped_old_rows": 0,
        "new_rows": 0,
        "new_by_ym": {},
        "qty_sum_by_month": {str(m): 0 for m in sorted(target_set)},
        "unmatched_country": [],
        "unmatched_series": [],
        "unmatched_oem": [],
        "matching_counts": {k: len(v) for k, v in matching.items()},
    }

    out_wb = Workbook(write_only=True)
    out_ws = out_wb.create_sheet("mkls-CV")
    out_ws.append(HEADER17)

    wb_old = openpyxl.load_workbook(main_path, read_only=True, data_only=True)
    old_ws = wb_old["mkls-CV"]
    old_ws.reset_dimensions()
    for row in old_ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        try:
            ym = (int(row[0]), int(row[1]))
        except (TypeError, ValueError):
            continue
        if ym in target_ym:
            stats["skipped_old_rows"] += 1
            continue
        out_ws.append(list(row[:17]))
        stats["kept_rows"] += 1
    wb_old.close()

    wb_raw = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)
    raw_ws = wb_raw[wb_raw.sheetnames[0]]
    raw_ws.reset_dimensions()
    header = next(raw_ws.iter_rows(min_row=2, max_row=2, values_only=True))
    target_cols = []
    for idx, value in enumerate(header):
        if idx < 7 or value is None:
            continue
        try:
            month = int(float(value))
        except (TypeError, ValueError):
            continue
        if month in target_set:
            target_cols.append((idx, month))
    if len(target_cols) != len(target_set):
        found = {m for _, m in target_cols}
        raise RuntimeError(f"missing target month columns: {sorted(target_set - found)}")

    unmatched_country = set()
    unmatched_series = set()
    unmatched_oem = set()
    new_by_ym = Counter()
    qty_sum_by_month = Counter()

    for row in raw_ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 7 or row[0] is None:
            continue
        country = str(row[0]).strip()
        if country == "中国":
            continue
        raw_vehicle = str(row[3]).strip() if row[3] else ""
        standard_vehicle = matching["vehicle_class"].get(raw_vehicle)
        if standard_vehicle != "轻型车":
            continue

        group = str(row[1]).strip() if row[1] else ""
        brand = str(row[2]).strip() if row[2] else ""
        segment = row[4]
        model = row[5]
        raw_power = str(row[6]).strip() if row[6] else ""
        standard_power = matching["power_map"].get(raw_power, "ICE")
        geo = matching["country_geo"].get(country)
        if geo is None:
            unmatched_country.add(country)
            geo = {"continent": None, "region": None, "new_region": None}
        series_key = matching["brand_series_idx"].get(brand) or matching["brand_series_idx"].get(strip_suffix(brand))
        series = matching["brand_series"].get(series_key) if series_key else None
        if series is None and brand:
            unmatched_series.add(brand)
        oem_key = matching["brand_oem_idx"].get(brand) or matching["brand_oem_idx"].get(strip_suffix(brand))
        oem = matching["brand_oem"].get(oem_key) if oem_key else None
        if oem is None and brand:
            unmatched_oem.add(brand)

        for idx, month in target_cols:
            if idx >= len(row):
                continue
            value = row[idx]
            if value is None:
                continue
            try:
                qty = float(value)
            except (TypeError, ValueError):
                continue
            if qty == 0:
                continue
            year, month_num = ym_from_int(month)
            qty_out = int(qty) if qty.is_integer() else qty
            out_ws.append(
                [
                    year,
                    month_num,
                    country,
                    group,
                    brand,
                    geo["continent"],
                    geo["region"],
                    "轻型车",
                    standard_power,
                    raw_vehicle,
                    segment,
                    model,
                    geo["new_region"],
                    model,
                    series,
                    oem,
                    qty_out,
                ]
            )
            stats["new_rows"] += 1
            new_by_ym[(year, month_num)] += 1
            qty_sum_by_month[month] += qty
    wb_raw.close()

    stats["new_by_ym"] = {f"{y}-{m:02d}": c for (y, m), c in sorted(new_by_ym.items())}
    stats["qty_sum_by_month"] = {str(k): int(v) if float(v).is_integer() else v for k, v in sorted(qty_sum_by_month.items())}
    stats["unmatched_country"] = sorted(unmatched_country)
    stats["unmatched_series"] = sorted(unmatched_series)
    stats["unmatched_oem"] = sorted(unmatched_oem)
    stats["total_rows_after"] = stats["kept_rows"] + stats["new_rows"]

    out_wb.save(out_path)
    return stats


COL_LETTERS = [chr(ord("A") + i) for i in range(17)]


def xml_escape(value):
    return str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def encode_row(row_num, values):
    parts = [f'<row r="{row_num}">']
    for col_idx, value in enumerate(values[:17]):
        if value is None:
            continue
        ref = COL_LETTERS[col_idx] + str(row_num)
        if isinstance(value, bool):
            parts.append(f'<c r="{ref}" t="b"><v>{1 if value else 0}</v></c>')
        elif isinstance(value, (int, float)):
            if isinstance(value, float) and value.is_integer():
                value = int(value)
            parts.append(f'<c r="{ref}"><v>{value}</v></c>')
        else:
            parts.append(
                f'<c r="{ref}" t="inlineStr"><is><t xml:space="preserve">{xml_escape(value)}</t></is></c>'
            )
    parts.append("</row>")
    return "".join(parts).encode("utf-8")


def replace_mkls_sheet(main_path, merged_path):
    backup = main_path.replace(
        ".xlsx", f"_preMklsBackup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    shutil.copy2(main_path, backup)

    wb = openpyxl.load_workbook(merged_path, read_only=True, data_only=True)
    ws = wb["mkls-CV"]
    ws.reset_dimensions()
    total_rows = sum(1 for _ in ws.iter_rows(values_only=True))
    wb.close()

    sheet_path = "xl/worksheets/sheet4.xml"
    tmp = main_path + ".mkls.tmp"
    sheet_open = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
        'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac" '
        'mc:Ignorable="x14ac">'
        f'<dimension ref="A1:Q{total_rows}"/>'
        '<sheetViews><sheetView workbookViewId="0"/></sheetViews>'
        '<sheetFormatPr defaultRowHeight="15" x14ac:dyDescent="0.25"/>'
        "<sheetData>"
    ).encode("utf-8")
    sheet_close = b"</sheetData></worksheet>"

    with zipfile.ZipFile(main_path, "r") as zin, zipfile.ZipFile(
        tmp, "w", zipfile.ZIP_DEFLATED, compresslevel=6, allowZip64=True
    ) as zout:
        for item in zin.infolist():
            if item.filename == sheet_path:
                with zout.open(item, "w", force_zip64=True) as fp:
                    fp.write(sheet_open)
                    fp.write(encode_row(1, HEADER17))
                    wb = openpyxl.load_workbook(merged_path, read_only=True, data_only=True)
                    ws = wb["mkls-CV"]
                    ws.reset_dimensions()
                    row_num = 2
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or all(v is None for v in row[:17]):
                            continue
                        fp.write(encode_row(row_num, row))
                        row_num += 1
                    wb.close()
                    fp.write(sheet_close)
            else:
                zout.writestr(item, zin.read(item.filename))
    os.replace(tmp, main_path)
    return {"backup": backup, "total_rows_with_header": total_rows}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--main", required=True)
    parser.add_argument("--raw", required=True)
    parser.add_argument("--merged", required=True)
    parser.add_argument("--report", required=True)
    parser.add_argument("--months", nargs="+", required=True, type=int)
    parser.add_argument("--writeback", action="store_true")
    args = parser.parse_args()

    t0 = time.time()
    stats = build_merged(args.main, args.raw, args.merged, args.months)
    stats["elapsed_build_sec"] = round(time.time() - t0, 1)
    stats["status"] = "validated"

    if stats["unmatched_country"] or stats["unmatched_series"] or stats["unmatched_oem"]:
        stats["status"] = "blocked_unmatched"
    elif args.writeback:
        stats["writeback"] = replace_mkls_sheet(args.main, args.merged)
        stats["status"] = "written"

    with open(args.report, "w", encoding="utf-8") as fp:
        json.dump(stats, fp, ensure_ascii=False, indent=2)
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    if stats["status"] == "blocked_unmatched":
        sys.exit(2)


if __name__ == "__main__":
    main()
