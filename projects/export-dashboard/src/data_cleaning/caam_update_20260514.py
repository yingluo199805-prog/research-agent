import argparse
import json
import os
import re
import shutil
import sys
import time
from collections import Counter
from datetime import datetime

import openpyxl


def strip_suffix(name):
    return re.sub(r"\s*\(.*?\)\s*", "", str(name)).strip()


def load_matching(main_path):
    wb = openpyxl.load_workbook(main_path, read_only=True, data_only=True)
    ws = wb["中汽协匹配字段"]
    ws.reset_dimensions()

    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    col_brand_series = col_brand_oem = col_multi = None
    for i, h in enumerate(row1):
        if h is None:
            continue
        s = str(h)
        if "品牌-系别" in s:
            col_brand_series = i + 1
        if "品牌-车企划分" in s:
            col_brand_oem = i + 1
        if "一对多" in s:
            col_multi = i + 1

    missing = [
        name
        for name, col in [
            ("品牌-系别", col_brand_series),
            ("品牌-车企划分", col_brand_oem),
            ("一对多", col_multi),
        ]
        if col is None
    ]
    if missing:
        raise RuntimeError(f"missing matching columns: {missing}")

    brand_series = {}
    for row in ws.iter_rows(
        min_row=3, min_col=col_brand_series, max_col=col_brand_series + 1, values_only=True
    ):
        if row[0] and row[1]:
            brand_series[str(row[0]).strip()] = str(row[1]).strip()
    bs_idx = {}
    for k in brand_series:
        bs_idx[k] = k
        bs_idx[strip_suffix(k)] = k

    brand_oem = {}
    for row in ws.iter_rows(
        min_row=3, min_col=col_brand_oem, max_col=col_brand_oem + 1, values_only=True
    ):
        if row[0] and row[1]:
            brand_oem[str(row[0]).strip()] = str(row[1]).strip()
    bo_idx = {}
    for k in brand_oem:
        bo_idx[k] = k
        bo_idx[strip_suffix(k)] = k

    multi_brand_set = set()
    multi_brand_factory = {}
    for row in ws.iter_rows(min_row=3, min_col=col_multi, max_col=col_multi + 3, values_only=True):
        if row[0] and row[1]:
            b = str(row[0]).strip()
            oem_div = str(row[1]).strip()
            f = str(row[2]).strip() if row[2] else ""
            multi_brand_set.add(b)
            multi_brand_set.add(strip_suffix(b))
            if f:
                multi_brand_factory[(b, f)] = oem_div

    mf_idx = {}
    for (b, f), v in multi_brand_factory.items():
        mf_idx[(b, f)] = v
        mf_idx[(strip_suffix(b), strip_suffix(f))] = v
        mf_idx[(strip_suffix(b), f)] = v
        mf_idx[(b, strip_suffix(f))] = v

    ws_cv = wb["中汽协CV "]
    ws_cv.reset_dimensions()
    existing_ym = set()
    existing_rows = 0
    for row in ws_cv.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        if row[0] is None:
            continue
        existing_rows += 1
        try:
            existing_ym.add((int(row[0]), int(row[1])))
        except (TypeError, ValueError):
            continue

    wb.close()
    return {
        "brand_series": brand_series,
        "bs_idx": bs_idx,
        "brand_oem": brand_oem,
        "bo_idx": bo_idx,
        "multi_brand_set": multi_brand_set,
        "mf_idx": mf_idx,
        "existing_ym": existing_ym,
        "existing_rows": existing_rows,
        "stats": {
            "brand_series": len(brand_series),
            "brand_oem": len(brand_oem),
            "multi_brands": len(multi_brand_set),
            "factory_pairs": len(mf_idx),
            "existing_months": len(existing_ym),
            "existing_rows": existing_rows,
        },
    }


def convert_raw(raw_path, matching):
    wb_raw = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)
    target_sheets = [s for s in wb_raw.sheetnames if "按中国整车厂商与合资公司" in s]
    if not target_sheets:
        raise RuntimeError("no target raw sheet found")

    long_data = []
    filtered_total = 0
    filtered_brand_na = 0
    sheet_rows = {}
    existing_ym = matching["existing_ym"]

    for sname in target_sheets:
        ws = wb_raw[sname]
        ws.reset_dimensions()
        m = re.search(r"(\d{4})", sname)
        if not m:
            continue
        year = int(m.group(1))
        rows = ws.iter_rows(values_only=True)
        next(rows, None)
        next(rows, None)
        header = list(next(rows))
        month_cols = {}
        for idx, h in enumerate(header):
            if h is None:
                continue
            mt = re.match(r"^(\d{1,2})月", str(h).strip())
            if mt:
                month_cols[int(mt.group(1))] = idx

        added = 0
        for row in rows:
            if row is None:
                continue
            vals = list(row)
            group = str(vals[0]).strip() if vals[0] else ""
            oem_factory = str(vals[1]).strip() if len(vals) > 1 and vals[1] else ""
            brand = str(vals[2]).strip() if len(vals) > 2 and vals[2] else ""
            model = str(vals[3]).strip() if len(vals) > 3 and vals[3] else ""

            if "Total" in group or "Total" in oem_factory:
                filtered_total += 1
                continue
            if not brand or brand in ("None", "n.a."):
                filtered_brand_na += 1
                continue

            for month_num, col_idx in month_cols.items():
                qty = vals[col_idx] if col_idx < len(vals) else None
                if qty is None:
                    continue
                if (year, month_num) in existing_ym:
                    continue
                long_data.append([year, month_num, f"{month_num}月", group, oem_factory, brand, model, qty])
                added += 1
        sheet_rows[sname] = added

    wb_raw.close()
    return long_data, {
        "target_sheets": target_sheets,
        "sheet_rows": sheet_rows,
        "filtered_total": filtered_total,
        "filtered_brand_na": filtered_brand_na,
    }


def match_rows(long_data, matching):
    cleaned = []
    unmatched_brand_series = set()
    unmatched_brand_oem = set()
    unmatched_factory = []

    for row in long_data:
        year, month_num, month_label, group, oem_factory, brand, model, qty = row

        key = matching["bs_idx"].get(brand) or matching["bs_idx"].get(strip_suffix(brand))
        series = matching["brand_series"][key] if key else "未匹配"
        if series == "未匹配":
            unmatched_brand_series.add(brand)

        if brand in matching["multi_brand_set"] or strip_suffix(brand) in matching["multi_brand_set"]:
            oem_division = (
                matching["mf_idx"].get((brand, oem_factory))
                or matching["mf_idx"].get((strip_suffix(brand), strip_suffix(oem_factory)))
                or matching["mf_idx"].get((strip_suffix(brand), oem_factory))
                or matching["mf_idx"].get((brand, strip_suffix(oem_factory)))
                or "未匹配"
            )
            if oem_division == "未匹配":
                unmatched_factory.append((brand, oem_factory))
        else:
            key2 = matching["bo_idx"].get(brand) or matching["bo_idx"].get(strip_suffix(brand))
            oem_division = matching["brand_oem"][key2] if key2 else "未匹配"
            if oem_division == "未匹配":
                unmatched_brand_oem.add(brand)

        cleaned.append([year, month_num, month_label, group, oem_factory, brand, model, qty, series, oem_division])

    return cleaned, {
        "unmatched_brand_series": sorted(unmatched_brand_series),
        "unmatched_brand_oem": sorted(unmatched_brand_oem),
        "unmatched_factory": sorted(set(unmatched_factory)),
    }


def append_cleaned(main_path, cleaned):
    backup_path = main_path.replace(
        ".xlsx", f"_preCaamBackup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    shutil.copy2(main_path, backup_path)

    wb = openpyxl.load_workbook(main_path)
    ws = wb["中汽协CV "]
    start_row = ws.max_row + 1
    for i, row in enumerate(cleaned):
        for j, val in enumerate(row):
            ws.cell(row=start_row + i, column=j + 1, value=val)
    wb.save(main_path)
    wb.close()
    return backup_path, start_row, start_row + len(cleaned) - 1


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--main", required=True)
    parser.add_argument("--raw", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--append", action="store_true")
    args = parser.parse_args()

    t0 = time.time()
    matching = load_matching(args.main)
    long_data, raw_stats = convert_raw(args.raw, matching)
    cleaned, unmatched = match_rows(long_data, matching)
    ym_counter = Counter((r[0], r[1]) for r in cleaned)
    brand_count = len(set(r[5] for r in cleaned))
    oem_count = len(set(r[9] for r in cleaned))

    payload = {
        "main": args.main,
        "raw": args.raw,
        "matching_stats": matching["stats"],
        "raw_stats": raw_stats,
        "cleaned_rows": len(cleaned),
        "year_month_rows": {f"{y}-{m:02d}": n for (y, m), n in sorted(ym_counter.items())},
        "brand_count": brand_count,
        "oem_count": oem_count,
        "unmatched": unmatched,
        "cleaned": cleaned,
        "elapsed_sec": round(time.time() - t0, 1),
    }

    if any(unmatched.values()):
        payload["status"] = "blocked_unmatched"
    elif args.append:
        backup_path, start_row, end_row = append_cleaned(args.main, cleaned)
        payload["status"] = "appended"
        payload["backup_path"] = backup_path
        payload["append_range"] = [start_row, end_row]
    else:
        payload["status"] = "validated"

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as fp:
        json.dump(payload, fp, ensure_ascii=False, indent=2, default=str)

    print(json.dumps({k: v for k, v in payload.items() if k != "cleaned"}, ensure_ascii=False, indent=2))
    if payload["status"] == "blocked_unmatched":
        sys.exit(2)


if __name__ == "__main__":
    main()
