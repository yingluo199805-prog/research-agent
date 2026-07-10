# -*- coding: utf-8 -*-
"""Relayout tables to user column order; copy summary template formatting."""

from __future__ import annotations

import argparse
import shutil
from pathlib import Path
from typing import Any

import openpyxl
import pythoncom
import win32com.client as win32

from m15_layout import (
    BLOCK_STRIDE,
    BLOCK_WIDTH,
    SALES_HEADERS,
    SHARE_HEADERS,
    is_share_header_row,
    m15_col_after_inserts,
    normalize_header,
    target_headers,
)
from update_workbook_m15_inline_fast import SHEET_TO_REGION, cell_addr, col_letter, find_block_start, text


XL_CALC_AUTOMATIC = -4105
XL_PASTE_ALL = -4104
XL_PASTE_FORMATS = -4122

ALL_SHEETS = list(SHEET_TO_REGION.keys())


def collect_qcols(path: Path, sheet_name: str) -> list[int]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    qcols = sorted(
        {
            cell.column
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row)
            for cell in row
            if cell.value == "2026Q1"
        }
    )
    wb.close()
    return qcols


def table_end_row(ws: Any, header_row: int, block_start: int, block_end: int, qcol: int) -> int:
    row = header_row + 1
    max_row = int(ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1)
    while row <= max_row + 5:
        label = text(ws.Cells(row, block_start).Value)
        if text(ws.Cells(row, qcol).Value) == "2026Q1":
            break
        if label.startswith(("一、", "二、", "三、", "四、", "五、", "六、", "七、", "八、", "九、", "十、")):
            break
        blank = True
        for col in range(block_start, block_end + 1):
            if text(ws.Cells(row, col).Value):
                blank = False
                break
        if blank and row > header_row + 1:
            break
        row += 1
    return row - 1


def discover_header_rows(ws: Any) -> list[tuple[int, int, bool]]:
    """Return (header_row, block_start, share) for each table."""
    items: list[tuple[int, int, bool]] = []
    seen: set[tuple[int, int]] = set()
    max_row = int(ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1)
    max_col = int(ws.UsedRange.Column + ws.UsedRange.Columns.Count - 1)

    block_starts: set[int] = {1}
    for col in range(1, max_col + 1):
        val = text(ws.Cells(1, col).Value)
        if val and col > 1 and text(ws.Cells(1, col - 1).Value) == "":
            block_starts.add(col)
    for col in range(1, max_col + 1):
        val = text(ws.Cells(3, col).Value)
        if val and val.startswith("一、"):
            block_starts.add(col)

    for block_start in sorted(block_starts):
        block_end = block_start + BLOCK_WIDTH - 1
        for header_row in range(1, max_row + 1):
            qcol = None
            for c in range(block_start + 1, min(block_end, max_col) + 1):
                if text(ws.Cells(header_row, c).Value) == "2026Q1":
                    qcol = c
                    break
            if qcol is None:
                continue
            if (header_row, block_start) in seen:
                continue
            seen.add((header_row, block_start))
            hdrs: dict[int, str | None] = {}
            for c in range(block_start + 1, block_end + 1):
                hdrs[c] = normalize_header(ws.Cells(header_row, c).Value)
            share = is_share_header_row(hdrs)
            items.append((header_row, block_start, share))
    return items


def find_block_start_openpyxl_ws(ws: Any, header_row: int, qcol: int) -> int:
    col = qcol
    while col > 1:
        if text(ws.Cells(header_row, col - 1).Value) == "":
            return col
        col -= 1
    return 1


def std_q1_col(block_start: int) -> int:
    return block_start + 5


def std_m15_col(block_start: int) -> int:
    return block_start + 6


def read_cell(cell: Any) -> Any:
    """Preserve formulas when remapping columns."""
    try:
        formula = cell.Formula
    except Exception:
        formula = ""
    if formula and isinstance(formula, str) and formula.startswith("="):
        return formula
    return cell.Value


def relayout_table(ws: Any, header_row: int, block_start: int, share: bool) -> None:
    block_end = block_start + BLOCK_WIDTH - 1
    targets = target_headers(share)
    qcol = std_q1_col(block_start)
    end_row = table_end_row(ws, header_row, block_start, block_end, qcol)

    header_map: dict[str, int] = {}
    for col in range(block_start + 1, block_end + 1):
        key = normalize_header(ws.Cells(header_row, col).Value)
        if key:
            header_map[key] = col

    matrix: list[list[Any]] = []
    label_col = block_start
    matrix.append([ws.Cells(header_row, label_col).Value] + targets)
    for row in range(header_row + 1, end_row + 1):
        label = ws.Cells(row, label_col).Value
        row_map: dict[str, Any] = {}
        for key, col in header_map.items():
            row_map[key] = read_cell(ws.Cells(row, col))
        matrix.append([label] + [row_map.get(h) if h else None for h in targets])

    address = f"{col_letter(block_start)}{header_row}:{col_letter(block_end)}{end_row}"
    ws.Range(address).Value = tuple(tuple(r) for r in matrix)

    m15_col = std_m15_col(block_start)
    if share:
        chg_col = block_start + 12
        for row in range(header_row + 1, end_row + 1):
            ws.Cells(row, chg_col).Formula = f"={cell_addr(row, m15_col + 1)}-{cell_addr(row, m15_col)}"
    else:
        delta_col = block_start + 12
        yoy_col = block_start + 13
        for row in range(header_row + 1, end_row + 1):
            ws.Cells(row, delta_col).Formula = f"={cell_addr(row, m15_col + 1)}-{cell_addr(row, m15_col)}"
            ws.Cells(row, yoy_col).Formula = (
                f'=IFERROR({cell_addr(row, m15_col + 1)}/{cell_addr(row, m15_col)}-1,"")'
            )


def apply_block_format(ws: Any, template_ws: Any, block_start: int, max_row: int) -> None:
    end_col = block_start + BLOCK_WIDTH - 1
    template_ws.Range(f"A1:{col_letter(BLOCK_WIDTH)}{max_row}").Copy()
    ws.Range(f"{col_letter(block_start)}1:{col_letter(end_col)}{max_row}").PasteSpecial(Paste=XL_PASTE_FORMATS)


def copy_summary_values(template_ws: Any, target_ws: Any, max_row: int = 175) -> None:
    template_ws.Range(f"A1:N{max_row}").Copy()
    target_ws.Range(f"A1:N{max_row}").PasteSpecial(Paste=XL_PASTE_ALL)


def relayout_sheet(
    workbook_path: Path,
    sheet_name: str,
    template_path: Path | None,
    preserve_summary: bool,
    max_row: int = 175,
) -> None:
    pythoncom.CoInitialize()
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    try:
        wb = excel.Workbooks.Open(str(workbook_path.resolve()))
        ws = wb.Worksheets(sheet_name)
        template_ws = None
        if template_path:
            t_wb = excel.Workbooks.Open(str(template_path.resolve()), ReadOnly=True)
            template_ws = t_wb.Worksheets("欧洲-核心数据")

        tables = discover_header_rows(ws)
        for header_row, block_start, share in tables:
            if preserve_summary and block_start == 1:
                continue
            relayout_table(ws, header_row, block_start, share)

        if template_ws is not None:
            if preserve_summary:
                copy_summary_values(template_ws, ws, max_row)
            block_starts = sorted({bs for _, bs, _ in tables})
            for bs in block_starts:
                apply_block_format(ws, template_ws, bs, max_row)

        excel.CutCopyMode = False
        excel.Calculation = XL_CALC_AUTOMATIC
        wb.Save()
        wb.Close(SaveChanges=True)
        if template_path:
            t_wb.Close(SaveChanges=False)
    finally:
        excel.ScreenUpdating = True
        excel.DisplayAlerts = True
        excel.Quit()
        pythoncom.CoUninitialize()


def run_full(
    cache_path: Path,
    base_path: Path,
    template_path: Path,
    output_path: Path,
) -> None:
    import json
    from update_workbook_m15_fixed import update_sheet

    cache = json.loads(cache_path.read_text(encoding="utf-8"))

    if output_path.exists():
        output_path.unlink()
    shutil.copy2(base_path, output_path)

    for sheet in ALL_SHEETS:
        src = template_path if sheet == "欧洲-核心数据" else base_path
        if sheet == "欧洲-核心数据":
            if output_path.exists():
                output_path.unlink()
            shutil.copy2(template_path, output_path)
            relayout_sheet(output_path, sheet, template_path, preserve_summary=True)
        else:
            update_sheet(cache, base_path, output_path, sheet, skip_if_m15=False)
            relayout_sheet(output_path, sheet, template_path, preserve_summary=False)
        print(f"completed {sheet}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--template", type=Path, default=None)
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--preserve-summary", action="store_true")
    parser.add_argument("--max-row", type=int, default=175)
    args = parser.parse_args()
    relayout_sheet(args.workbook, args.sheet, args.template, args.preserve_summary, args.max_row)
    print(f"relayout done: {args.sheet}")


if __name__ == "__main__":
    main()
