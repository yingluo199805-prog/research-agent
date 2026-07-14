# -*- coding: utf-8 -*-
"""
底稿生成器 —— 把年报中"数据出处"的页面截屏，按年份排序嵌入一个 Excel。

底稿的用途是"方便检查"：审表的人不必翻 PDF，在一个 Excel 里逐年看到
原始附注页的图像，旁边就是页码出处。所以底稿存的是**页面截图**，不是文本。

用法：
    python build_digao.py spec.json

spec.json 结构（页码为 PDF 物理页码，1 开始；可写连续多页）：
{
  "company": "长城汽车",
  "output": "C:/.../长城汽车_年报附注底稿.xlsx",
  "dpi": 150,                         // 可选，默认 150
  "subjects": [
    {
      "name": "政府补助分析",          // 一个科目 = 一个 sheet
      "pages": [
        {"year": 2018, "pdf": "C:/.../GWM_2018.pdf", "note": "附注六/十一", "page_numbers": [205, 206]},
        {"year": 2019, "pdf": "C:/.../GWM_2019.pdf", "note": "附注六/十一", "page_numbers": [210]}
      ]
    }
  ]
}

依赖：PyMuPDF(fitz)、openpyxl、Pillow。脚本不联网，纯本地渲染。
"""
import sys, os, json, math
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TARGET_W = 1000          # 嵌入图像目标宽度(px)，过大则文件臃肿、过小看不清字
PX_PER_ROW = 18          # 估算图像占多少行用于推进锚点(每行约18px)
GAP_ROWS = 3             # 每张图之间留白行
HEADER_BLUE = "1F4E78"


def render_page(doc, pno_1based, dpi, out_png):
    """渲染单页为 PNG，返回 (宽px, 高px)。"""
    page = doc[pno_1based - 1]
    zoom = dpi / 72.0
    pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
    pix.save(out_png)
    return pix.width, pix.height


def build(spec_path):
    with open(spec_path, encoding="utf-8") as f:
        spec = json.load(f)

    company = spec.get("company", "")
    output = spec["output"]
    dpi = spec.get("dpi", 150)
    img_dir = os.path.join(os.path.dirname(output), "_digao_imgs")
    os.makedirs(img_dir, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    doc_cache = {}
    img_seq = 0

    for subj in spec["subjects"]:
        sname = subj["name"]
        ws = wb.create_sheet(title=sname[:31])
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 150

        # 标题区
        ws["B1"] = f"{company} {sname} 年报附注底稿 —— 数据出处页面截图（按年份升序）"
        ws["B1"].font = Font(name="Microsoft YaHei", size=13, bold=True, color=HEADER_BLUE)
        ws["B2"] = "用途：逐年核对建模数据与年报原文。每段为该年附注所在页的整页截图，标题行注明年报/附注/页码。"
        ws["B2"].font = Font(name="Microsoft YaHei", size=9, italic=True)

        row = 4
        # 按年份升序排，方便检查
        for entry in sorted(subj["pages"], key=lambda e: e["year"]):
            year = entry["year"]
            pdf = entry["pdf"]
            note = entry.get("note", "")
            pages = entry["page_numbers"]
            if pdf not in doc_cache:
                doc_cache[pdf] = fitz.open(pdf)
            doc = doc_cache[pdf]

            # 年份分隔标题行
            label = (f"▎{year} 年报 — {note} — {os.path.basename(pdf)} — "
                     f"第 {','.join(str(p) for p in pages)} 页")
            cell = ws.cell(row=row, column=2, value=label)
            cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=HEADER_BLUE)
            cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[row].height = 22
            row += 1

            for pno in pages:
                img_seq += 1
                png = os.path.join(img_dir, f"{sname}_{year}_{pno}_{img_seq}.png")
                w, h = render_page(doc, pno, dpi, png)
                scale = TARGET_W / w
                disp_w = int(w * scale)
                disp_h = int(h * scale)
                xim = XLImage(png)
                xim.width = disp_w
                xim.height = disp_h
                anchor = f"B{row}"
                ws.add_image(xim, anchor)
                rows_used = math.ceil(disp_h / PX_PER_ROW)
                row += rows_used + GAP_ROWS

    for d in doc_cache.values():
        d.close()

    wb.save(output)
    print(json.dumps({
        "status": "ok",
        "output": output,
        "sheets": [s for s in wb.sheetnames],
        "images_embedded": img_seq
    }, ensure_ascii=False))


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("usage: python build_digao.py spec.json", file=sys.stderr)
        sys.exit(1)
    build(sys.argv[1])
